#!/usr/bin/env python
# coding=utf-8
"""
Script to save city object instance data to txt file for better overview.

Currently, only building entities (not PV- or Wind-farms) are included.
Output file format is close
"""
from __future__ import division

import os
import pickle
import warnings
import numpy as np

try:
    import openpyxl
except:
    msg = 'Could not import openpyxl. Which is required, if you want to' \
          ' save profiles directly into xlsx files. Please install via ' \
          'pip or set save_as_xlsx to False.'
    warnings.warn(msg)


def save_city_data_to_file(city, save_path, with_esys=False,
                           use_german=False, save_as_xlsx=False):
    """
    Saves city object data to output file.

    Parameters
    ----------
    city : object
        City object of PyCity
    save_path : str
        Path to save data to
    with_esys : bool, optional
        Defines, if information about energy systems should be saved, too.
        (default: False)
    use_german : bool, optional
        Defines, if English or German language should be used
        (default: False). If False, uses English language.
    save_as_xlsx : bool, optional
        Define, if load curves should also be saved as xlsx files
        (default: False)
    """

    if with_esys is False:
        nb_params = 23  # Until construction type
    else:
        #  TODO: To be added
        msg = 'Extraction of energy system data has NOT been implemented, yet!'
        warnings.warn(msg)

    # Get number of buildings
    nb_build = 0
    for n in city:
        #  If node holds attribute 'node_type'
        if 'node_type' in city.nodes[n]:
            #  If node_type is building
            if city.nodes[n]['node_type'] == 'building':
                #  If entity is kind building
                if city.nodes[n]['entity']._kind == 'building':
                    nb_build += 1

    # Define 2d zeros array (to be filled with data)
    data_array = np.zeros((nb_build, nb_params)) * np.nan

    #  Initial row
    row = 0

    for n in city:
        #  If node holds attribute 'node_type'
        if 'node_type' in city.nodes[n]:
            #  If node_type is building
            if city.nodes[n]['node_type'] == 'building':

                if 'entity' in city.nodes[n]:
                    #  If entity is kind building
                    if city.nodes[n]['entity']._kind == 'building':

                        #  Define pointer to current building
                        cur_b = city.nodes[n]['entity']

                        #  Write data into data_array

                        #  id
                        data_array[row][0] = n

                        #  x coordinate
                        data_array[row][1] = city.nodes[n]['position'].x

                        #  y coordinate
                        data_array[row][2] = city.nodes[n]['position'].y

                        #  build_type
                        data_array[row][3] = cur_b.build_type

                        #  net floor area
                        data_array[row][4] = cur_b.net_floor_area

                        #  build year
                        data_array[row][5] = cur_b.build_year

                        #  mod year
                        data_array[row][6] = cur_b.mod_year

                        #  Annual space heating energy demand
                        data_array[row][7] = cur_b.get_annual_space_heat_demand()

                        #  Annual el. heating energy demand
                        data_array[row][8] = cur_b.get_annual_el_demand()

                        #  Usable PV roof
                        data_array[row][9] = cur_b.roof_usabl_pv_area

                        #  Number of apartments
                        data_array[row][10] = cur_b.get_number_of_apartments()

                        #  Number of occupants
                        data_array[row][11] = cur_b.get_number_of_occupants()

                        #  Number of floors
                        data_array[row][12] = cur_b.nb_of_floors

                        #  Height of floors
                        data_array[row][13] = cur_b.height_of_floors

                        #  AHU
                        data_array[row][14] = cur_b.with_ahu

                        #  residential_layout
                        data_array[row][15] = cur_b.residential_layout

                        #  neighbour_buildings
                        data_array[row][16] = cur_b.neighbour_buildings

                        #  attic
                        data_array[row][17] = cur_b.attic

                        #  cellar
                        data_array[row][18] = cur_b.cellar

                        #  dormer
                        data_array[row][19] = cur_b.dormer

                        #  construction_type
                        if cur_b.construction_type == 'heavy':
                            constr_id = 0
                        elif cur_b.construction_type == 'light':
                            constr_id = 1
                        else:
                            constr_id = 0

                        data_array[row][20] = constr_id

                        #  Annual hot water heating energy demand
                        data_array[row][21] = cur_b.get_annual_dhw_demand()

                        #  Ground area
                        data_array[row][22] = cur_b.ground_area

                        #  Go one row down
                        row += 1

                    elif city.nodes[n]['entity']._kind == 'windenergyconverter':
                        #  TODO: To be implemented (plus new version of nb_build)
                        pass
                    elif city.nodes[n]['entity']._kind == 'pv':
                        #  TODO: To be implemented (plus new version of nb_build)
                        pass

    # Define header
    if use_german:
        header = u'ID\tX-Koordinate in m\tY-Koordinate in m\tGebäudetyp' \
                 u'\tNettogrundfläche in m2' \
                 u'\tBaujahr\tLetztes Sanierungsjahr' \
                 u'\tJährlicher Nutzenergiebedarf Raumwärme in kWh' \
                 u'\tJährlicher elektrischer Energiebedarf (ohne Warmwasser) in kWh' \
                 u'\tNutzbare Photovoltaik Dachfläche in m2' \
                 u'\tAnzahl Apartments\tAnzahl Bewohner' \
                 u'\tAnzahl Geschosse\tMittlere Geschosshöhe in m\tMit Klimaanlage' \
                 u'\tWohngebäudelayout\tNachbargebäude\tTyp Dach\tTyp Keller' \
                 u'\tDachgauben\tArt der Konstruktion\tWarmwasserbedarf in kWh\t' \
                 u'Bebaute Grundfläche in m2'

    else:
        header = 'id\tX\tY\tbuilding_type\ttab_ease_building_net_floor_area' \
                 '\ttab_ease_building_build_year\ttab_ease_building_mod_year' \
                 '\tAnnual space heat e demand in kWh' \
                 '\tAnnual electr. E demand in kWh\tUsable pv roof area in m2' \
                 '\tNumber of apartments\tTotal number of occupants' \
                 '\tNumber of floors\tHeight of floors\twith ahu' \
                 '\tresidential layout\tneighbour buildings\tattic\tcellar' \
                 '\tdormer\tconstruction type\tdhw e demand in kWh\t' \
                 'ground area in m2'

    #  Replace all None with np.nan to prevent saving errors
    for i in range(len(data_array)):
        for j in range(len(data_array[0])):
            if data_array[i][j] == None:
                data_array[i][j] = np.nan

    # Save to path
    np.savetxt(save_path, data_array, delimiter='\t', header=header)

    if save_as_xlsx:
        #  Get workbook
        wb = openpyxl.Workbook()

        #  Get worksheet
        ws = wb.active

        if use_german:
            ws['A1'].value = 'ID'
            ws['B1'].value = u'X-Koordinate in m'
            ws['C1'].value = u'Y-Koordinate in m'
            ws['D1'].value = u'Gebäudetyp'
            ws['E1'].value = U'Nettogrundfläche in m2'
            ws['F1'].value = u'Baujahr'
            ws['G1'].value = u'Letztes Sanierungsjahr'
            ws['H1'].value = u'Jährlicher Nutzenergiebedarf Raumwärme in kWh'
            ws['I1'].value = u'Jährlicher elektrischer Energiebedarf (ohne Warmwasser) in kWh'
            ws['J1'].value = u'Nutzbare Photovoltaik Dachfläche in m2'
            ws['K1'].value = u'Anzahl Apartments'
            ws['L1'].value = u'Anzahl Bewohner'
            ws['M1'].value = u'Anzahl Geschosse'
            ws['N1'].value = u'Mittlere Geschosshöhe in m'
            ws['O1'].value = u'Mit Klimaanlage'
            ws['P1'].value = u'Wohngebäudelayout'
            ws['Q1'].value = u'Nachbargebäude'
            ws['R1'].value = u'Typ Dach'
            ws['S1'].value = u'Typ Keller'
            ws['T1'].value = u'Dachgauben'
            ws['U1'].value = u'Art der Konstruktion'
            ws['V1'].value = u'Warmwasserbedarf in kWh'
            ws['W1'].value = u'Bebaute Grundfläche'

        else:
            ws['A1'].value = 'id'
            ws['B1'].value = 'x-coordinate in m'
            ws['C1'].value = 'y-coordinate in m'
            ws['D1'].value = 'Building type number'
            ws['E1'].value = U'Net floor area in m2'
            ws['F1'].value = u'Year of construction'
            ws['G1'].value = u'Last year of modernization'
            ws['H1'].value = u'Annual net space heating demand in kWh'
            ws['I1'].value = u'Annual electric energy demand (without hot water) in kWh'
            ws['J1'].value = u'Usable PV-area in m2'
            ws['K1'].value = u'Number of apartments'
            ws['L1'].value = u'Number of occupants'
            ws['M1'].value = u'Number of floors'
            ws['N1'].value = u'Average floor height in m'
            ws['O1'].value = u'With air handling unit (AHU)'
            ws['P1'].value = u'Residential layout'
            ws['Q1'].value = u'Connected neighbour buildings'
            ws['R1'].value = u'Type of attic'
            ws['S1'].value = u'Type of cellar'
            ws['T1'].value = u'Dormer'
            ws['U1'].value = u'Construction type'
            ws['V1'].value = u'Net hot water demand in kWh'
            ws['W1'].value = u'Used ground floor area in m2'

        # Loop over columns
        for j in range(len(data_array[0])):
            #  Loop over rows
            for i in range(len(data_array)):
                ws.cell(row=i + 2, column=j + 1, value=data_array[i][j])

        xlsx_filename = os.path.basename(save_path)[:-4] + '.xlsx'

        workbook_path = os.path.join(os.path.dirname(save_path),
                                     xlsx_filename)

        wb.save(workbook_path)

    print('Saved district data to path ' + str(save_path))


def extract_and_save_building_load_profiles(city, save_path):
    """
    Extract all building load profiles (space heating, electrical, hot water)
    and save results to file.

    Parameters
    ----------
    city : object
        City object of pycity_calc
    save_path : str
        Path to save file to
    """

    nb_buildings = city.get_nb_of_building_entities()

    #  Nb. columns 1 + 3 * nb_buildings

    timestep = city.environment.timer.timeDiscretization

    nb_timesteps = 365 * 24 * 3600 / timestep
    print(nb_timesteps)

    time_array = np.arange(start=0, stop=nb_timesteps*timestep, step=timestep)

    list_processed_ids = []

    for n in city.nodes():

        #  If node holds attribute 'node_type'
        if 'node_type' in city.nodes[n]:

            #  If node_type is building
            if city.nodes[n]['node_type'] == 'building':

                if 'entity' in city.nodes[n]:

                    #  If entity is kind building
                    if city.nodes[n]['entity']._kind == 'building':

                        curr_b = city.nodes[n]['entity']

                        sh_curve = curr_b.get_space_heating_power_curve()
                        time_array = np.vstack((time_array, sh_curve))

                        el_curve = curr_b.get_electric_power_curve()
                        time_array = np.vstack((time_array, el_curve))

                        dhw_curve = curr_b.get_dhw_power_curve()
                        time_array = np.vstack((time_array, dhw_curve))

                        list_processed_ids.append(n)

    res_array = np.transpose(time_array)

    #  Generate header
    header = 'Time in seconds\t'

    for elem in list_processed_ids:
        header += 'id ' + str(elem) + '\t\t\t'

    header += '\n\t'
    for elem in list_processed_ids:
        header += 'Space heating in W\tEl. power in W\tHot water power in W\t'

    print(header)

    np.savetxt(save_path, res_array, delimiter='\t', header=header)



if __name__ == '__main__':
    #  This path
    this_path = os.path.dirname(os.path.abspath(__file__))
    src_path = os.path.dirname(os.path.dirname(this_path))

    #  City file name (to
    city_f_name = 'city_3_buildings_mixed.pkl'

    use_german = False
    with_esys = False
    save_as_xlsx = True

    #  Path to load city object
    city_path = os.path.join(this_path, 'input', city_f_name)
    # city_path = os.path.join(src_path, 'cities', 'scripts', 'input_osm',
    #                          'Diss_Quartiere',
    #                          city_f_name)

    #  Txt output filename
    out_name = city_f_name[:-4] + '.txt'

    #  Path to save data to
    out_path = os.path.join(this_path, 'output', out_name)
    # out_path = os.path.join(src_path, 'cities', 'scripts', 'input_osm',
    #                          'Diss_Quartiere', out_name)

    #  Load city object
    city = pickle.load(open(city_path, mode='rb'))

    #  Extract and save city object data
    save_city_data_to_file(city=city, save_path=out_path, with_esys=with_esys,
                           use_german=use_german, save_as_xlsx=save_as_xlsx)

    #  Extract all building load profiles and write them to file
    #  extract_and_save_building_load_profiles(city=city, save_path=out_path)
